from task8.forms import MyForm
from django.shortcuts import render
from openpyxl import load_workbook


def add_patient(req):
    if req.method == 'POST':
        form = MyForm(req.POST)
        if form.is_valid():
            cd = form.cleaned_data
            wb = load_workbook('patients.xlsx')
            ws = wb['emp']
            num_rows = str(int(ws['F1'].value))
            ws['A'+num_rows] = cd['name']
            ws['B'+num_rows] = cd['email']
            ws['C'+num_rows] = cd['phone']
            ws['F1'] = int(num_rows) + 1
            wb.save('patients.xlsx')
            return render(req , 'patientregister.html',{"form" : form , 'msg' : "patient is registered"})
    else:
        form = MyForm()
        return render(req , 'patientregister.html',{"form" : form})