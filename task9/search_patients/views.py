import json
from search_patients.forms import MyForm
from django.http import HttpResponse
from openpyxl import load_workbook
from django.shortcuts import render

def search(req):
    if req.method == 'POST':
        form = MyForm(req.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            wb = load_workbook('patients.xlsx')
            ws = wb['emp']
            result = []
            for row in range(2 , ws.max_row + 1):
                p_name = ws.cell(row , 1).value
                if p_name.__contains__(name):
                    print("found a match")
                    # found a match
                    p = {}
                    p['name'] = ws.cell(row,1).value
                    p['email'] = ws.cell(row,2).value
                    p['phone'] = ws.cell(row,3).value
                    result.append(p)
            if len(result) == 0:
                result.append({'error' : 'no patients found'})
            # return HttpResponse(json.dumps(result))
            return render(req , 'search.html' , {'form' : form , 'result' : result})
            
    else:
        form = MyForm()
        return render(req , 'search.html' , {'form' : form})