import json
from django.shortcuts import render
from openpyxl import load_workbook


def displayHomePage(req) :
    wb = load_workbook('patients.xlsx')
    ws = wb['emp']
    patiensts = []

    for row in range(2, ws.max_row + 1):
        patient = {}
        patient['name'] = ws.cell(row , 1).value
        patient['email'] = ws.cell(row , 2).value
        patient['phone'] = ws.cell(row , 3).value
        patiensts.append(patient)

    
    return render(req , 'index.html' , {'patients' : patiensts})